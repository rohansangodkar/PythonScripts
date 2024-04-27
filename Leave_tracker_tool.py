import os

from pip._internal import main as pipmain

pipmain(['install', 'pandas','openpyxl','xlrd','pywin32'])

import win32com.client as win32
import pandas as pd
import xlrd
import re
import copy
import itertools
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pandas._libs.tslibs.timestamps import Timestamp

Res_sheet = pd.read_excel(r"C:\Users\singhalr\OneDrive - ACI Worldwide Corp\Automation\Leave_tracker\ACI_RES_SHEET.xlsx")

path = r"C:\Users\singhalr\OneDrive - ACI Worldwide Corp\Automation\Leave_tracker\test"

files =  os.listdir(path)
md={}

Res = {}
Leave = {}

for index,row in Res_sheet.iterrows():
    id = row['EMP ID']
    if id not in Res:
        Res[id] = row['Resource Name']
        Leave[id] = 0

for file in files:
    if '.xlsx' not in file and '.xls' in file:
        print(file)
        fpath = os.path.join(path, file)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fpath)
        nfpath = os.path.exists(fpath+"x")
        if not nfpath:
            wb.SaveAs(fpath+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()
    else:
        fpath = os.path.join(path, file)
        sheet = pd.read_excel(fpath)


    calendar = {'JAN': 0,'FEB':1,'MAR':2,'APR':3,'MAY':4,'JUN':5,'JUL':6,'AUG':7,'SEP':8,'OCT':9,'NOV':10,'DEC':11}

    y=sheet.columns[8].split('-')[2]
    
    for index,row in sheet.iterrows():
        id = row['EMP ID']
        if id not in Res:
            continue
        else:
            if row['Task'] == 'Leave' or row['Task'] == 'Public Holiday':
                if int(row['Total']%8) == 0:
                    Leave[id]=(int(row['Total'])//8)
                else:
                    Leave[id]=(int(row['Total'])//8)
                    Leave[id]=Leave[id]+0.5

    mm = str(sheet.columns[7]).split('-')[1]+'-'+y
    md[mm]=list(Leave.values())
    md = {k: md[k] for k in sorted(md, key=lambda x: datetime.strptime(x, '%b-%Y'))}
month=list(md.keys())

df = pd.DataFrame(columns= ['EMP ID'] + ['EMP NAME'] + month)

while True:   # repeat until the try statement succeeds
    try:
        path = input('Please, Enter the path where you want leave tracker to be generated: ')
        excel_writer = pd.ExcelWriter(path+'\\Leave_tracker.xlsx',engine='openpyxl')
        break                             # exit the loop
    except IOError:
        input("Could not open file or incorrect path given ! Please close Report Excel or give the correct. Press Enter to retry.")
        break
        # restart the loop
    
df.to_excel(excel_writer, sheet_name='tracker', index = False)

wb = excel_writer.book
ws = excel_writer.sheets["tracker"]

row_data = 2
for r in Res:
        col_data = 1
        ws.cell(row=row_data, column= col_data, value = r)
        col_data = 2
        ws.cell(row=row_data, column= col_data, value = Res[r])
        row_data = row_data+1

col_data = 3 
row_data = 2
for j in range(len(month)):
        col_data = col_data + j
        row_data = 2
        for i in range(len(md[month[j]])):
                ws.cell(row=row_data, column= col_data, value = (md[month[j]][i]))
                row_data = row_data+1

excel_writer._save()