import os

from pip._internal import main as pipmain

pipmain(['install', 'pandas','openpyxl','xlrd','pywin32'])

import win32com.client as win32
import pandas as pd
import xlrd
import re
import copy
import itertools
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pandas._libs.tslibs.timestamps import Timestamp

script_name = "Invoice Generation Tool!!!"
top_border = "\u256D" + "\u2500" * (len(script_name) + 4) + "\u256E"
bottom_border = "\u2570" + "\u2500" * (len(script_name) + 4) + "\u256F"
print(top_border)
print("\u2502", script_name, " ","\u2502")
print(bottom_border)

calendar = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}

month_input= input("Enter the month and year (yyyy-mm): ")
try:
    user_date = datetime.strptime(month_input, '%Y-%m')
except ValueError:
    print("Invalid date")
    exit(1)
    
year,month=map(int,month_input.split('-'))

cal = []
first= date(year,month,1)

if month == 12:
    num_days=(date(year+1,1,1)-first).days
else:
    num_days=(date(year,month+1,1)-first).days

for day in range(num_days):
    curr=first+timedelta(days=day)

    if curr.weekday()>=5:
        cal.append(1)
    else:
        cal.append(0)

COM = 'COM resourcing sheet'
EXP = 'SOW_Exceptions'

limit = 0

while limit <3 : 
        print(' ')
        path = input('Please, Enter the folder path of timesheets: ')
        if path[0]==r"'" or path[0]==r'"':
            path=path[1:]
        if path[-1]==r"'" or path[-1]==r'"':
            path=path[:len(path)-1]
        if os.path.isdir(path):
            limit = 0
            break                            
        else:
            limit +=1
            input("Wrong timesheets file path. Try again!")

Found = False
EXP_Flag = False

while limit <3 : 
        print(" ")
        data_file = input('Please, Enter the path of COM Resource File: ')
        if data_file[0]==r"'" or data_file[0]==r'"':
            data_file=data_file[1:]
        if data_file[-1]==r"'" or data_file[-1]==r'"':
            data_file=data_file[:len(data_file)-1]
        if os.path.isfile(data_file):
            limit = 0
            data = pd.ExcelFile(data_file)
            sheet=data.sheet_names
            Found = False
            for a in sheet:
                if a.lower().strip() == COM.lower().strip():
                    data_A = pd.read_excel(data,sheet_name=a)
                    Found = True
                if a.lower().strip() == EXP.lower().strip():
                    exp = pd.read_excel(data,sheet_name=a)   
                    EXP_Flag = True  
            if Found:
                limit = 0
                Found= False
                break
            else:
                limit +=1
                print('COM resourcing sheet - tab name not found. Try again')                           
        else:
            limit +=1
            input("Wrong COM Resource file path. Try again!")

if limit == 3:
    print("Too many attempts. Try again!")
    exit(1)

files =  os.listdir(path)

for file in files:
    if '.xlsx' not in file and '.xls' in file:
        print(file)
        fpath = os.path.join(path, file)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(fpath)
        nfpath = os.path.exists(fpath+"x")
        if not nfpath:
            wb.SaveAs(fpath+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()                               #FileFormat = 56 is for .xls extension
        excel.Application.Quit()

allid = data_A['EMP ID'].unique().tolist()

files = os.listdir(path)
Empid = {}
o_Empid = {}
visited= {}
nofile =[]
for file in files:
    fpath = os.path.join(path, file)
    id = re.search(r'\d{5}',file)
    if id != None:
        id = int(id.group())
        if id not in Empid:
            Empid[id] = [[],[],[],[],[]]
            visited[id] = []
    else:
        nofile.append(file)
        continue
    
    if '.xlsx' in file:
        F = pd.read_excel(fpath)
    elif '.xls' in file and '.xlsx' not in file:
        continue
    else:
        nofile.append(file)
        continue
    print(file)
    F.fillna(0,inplace=True)
    record_len = 0

    if 'Task' not in F.columns:
        nofile.append(file)
        continue
    else:
        for i in F['Task']:
            if isinstance(i,str):
                record_len = record_len+1
        F.dropna(subset='Task',inplace=True)
        F.columns = F.iloc[0]
    
    dict= 0

    month_data = calendar[month]
    dict= 0
    for i in F.columns:
        if month_data +' 01' in str(i):
            dict = 0
            break
        elif month_data +' 08' in str(i):
            dict = 1
            break
        elif month_data +' 15' in str(i):
            dict = 2
            break
        elif month_data +' 22' in str(i):
            dict = 3
            break
        elif month_data +' 29' in str(i):
            dict = 4
            break
        else:
            h = str(cal[0])+str(cal[1])
            if month == 2 and year%4!= 0 and h !='10' and month_data +' 28' in str(i):
                dict = 4
                print(dict)
            else:
                dict = -1
    j = 0
    Month_flag = False
    if dict != -1:
        while j < 7:
            Leave = False
            i=0
            s = 0
            if int(F.iloc[0,(4+j)].split(' ')[2]) not in visited[id]:
                while i < record_len:
                    if "PTO-Paid Time Off" in F.iloc[(2+i),2] and month_data in F.iloc[0,(4+j)] and F.iloc[(2+i),(4+j)] != 0:
                        Leave = True
                        break
                    else:
                        if month_data in F.iloc[0,(4+j)]:
                            F.iloc[(2):,(4+j)].fillna(0,inplace = True)
                            s= s + (F.iloc[(2+i),(4+j)])
                            Month_flag = True
                        else:
                            Month_flag = False
                            break
                    i+=1
                if month_data in F.iloc[0,(4+j)]:
                    visited[id].append(int(F.iloc[0,(4+j)].split(' ')[2]))
                if s == 0 and (('Saturday'in F.iloc[0,(4+j)]) or ('Sunday' in F.iloc[0,(4+j)])) and month_data in F.iloc[0,(4+j)]:
                    Empid[id][dict].append('WO')
                    Month_flag = False
                if Month_flag and not Leave:
                    if s > 9:
                        if id not in o_Empid:
                            o_Empid[id]=[]
                    if isinstance(s,float):
                        if str(s).split('.')[1] == '0':
                            Empid[id][dict].append(int(str(s).split('.')[0]))
                        else:
                            Empid[id][dict].append(s)
                    else:
                        Empid[id][dict].append(s)
                if s != 0 and (('Saturday'in F.iloc[0,(4+j)]) or ('Sunday' in F.iloc[0,(4+j)])) and month_data in F.iloc[0,(4+j)]:
                    if id not in o_Empid:
                        o_Empid[id] = []
                if Leave:
                    Empid[id][dict].append('EL')
                    Leave = False
                    Month_flag = False
                j+=1
            else:
                j+=1
nodate={}

for key in Empid:
    Empid[key] = list(itertools.chain.from_iterable(Empid[key]))

for key in Empid:
    visited[key].sort()
    dup = []
    i = 0
    while i <len(cal):
        if (i+1) not in visited[key]:
            Empid[key].insert(i,0)
            if key not in nodate:
                nodate[key] = []
            nodate[key].append(i+1)
        i+=1
    i=0
    if key in o_Empid:
        while i < len(cal):
            if cal[i] == 1 and Empid[key][i] != 0:
                o_Empid[key].insert(i,Empid[key][i])
                dup.append('WO')
            elif cal[i] == 1 and Empid[key][i] == 0:
                o_Empid[key].insert(i,'WO')
                dup.append('WO')
            else:
                if isinstance(Empid[key][i],str):
                    o_Empid[key].insert(i,Empid[key][i])
                    dup.append(Empid[key][i])
                elif Empid[key][i] > 9:
                    o_Empid[key].insert(i,(Empid[key][i]-9))
                    dup.append(9)
                else:
                    o_Empid[key].insert(i,0)
                    dup.append(Empid[key][i])
            i+=1
        Empid[key] = dup
    else:
        while i < len(cal):
            if cal[i] == 1:
                Empid[key][i]='WO'
            i+=1

bill = []
nonbill = []
unbill = 'Buffer - Client Committed '
nbill = {}
ybill = {}
name ={}
Empsow = {}
status = {}
sow = {}
sowdays = {}
shadow={}
Rate = {}

for e in Empid:    
    for index,rows in data_A.iterrows():
        s = rows['Oracle  status']
        unbill = 'Buffer - Client Committed '
        if rows['EMP ID'] == e and  s== 'Billable ':
            name[e] = rows['Resource Name']
            Empsow[e] = rows['2023 SOW#']
            status[e] = s
            Rate[e] = rows['Rate']
            if rows['2023 SOW#'] not in sow:
                sow[rows['2023 SOW#']]=[]
                sowdays[rows['2023 SOW#']]=[]
            if e not in sow[rows['2023 SOW#']]:
                sow[rows['2023 SOW#']].append(e)
            break
        elif rows['EMP ID'] == e and  s.lower().strip()== unbill.lower().strip():
            name[e] = rows['Resource Name']
            Empsow[e] = rows['2023 SOW#']
            status[e] = s
            Rate[e] = rows['Rate']
            if e not in shadow:
                shadow[e]=[]
            break

if EXP_Flag:
    if len(exp) > 0:
        exp.dropna(subset=['EMP ID'],inplace=True)  
        for index,rows in exp.iterrows():
            if len(str(rows['EMP ID']).split('.')[0]) == 5 and str(rows['EMP ID']).split('.')[0][0] == '7':
                if rows['Oracle  status'] != unbill:
                    l = rows.tolist()
                    l[0]=int(str(l[0]).split('.')[0])
                    l[4]=Timestamp(l[4])
                    l[4]= l[4].to_pydatetime()
                    l[4]=l[4].strftime('%d-%m-%Y')
                    l[5]=Timestamp(l[5])
                    l[5]= l[5].to_pydatetime()
                    l[5]=l[5].strftime('%d-%m-%Y') 
                    bill.append(l)
                else:
                    l = rows.tolist()
                    l[0] = int(str(l[0]).split('.')[0])
                    l[4]=Timestamp(l[4])
                    l[4]= l[4].to_pydatetime()
                    l[4]=l[4].strftime('%d-%m-%Y')
                    l[5]=Timestamp(l[5])
                    l[5]= l[5].to_pydatetime()
                    l[5]=l[5].strftime('%d-%m-%Y') 
                    nonbill.append(l)
            else:
                print('incorrect record found', str(rows['EMP ID']).split('.')[0])

#dict to have index for both billable and non billable emloyees

        for a in range(len(nonbill)):
            if nonbill[a][0] not in nbill:
                nbill[nonbill[a][0]] = []
            nbill[nonbill[a][0]].append(a)
            a+=1
        for a in range(len(bill)):
            if bill[a][0] not in ybill:
                ybill[bill[a][0]]=[]
            ybill[bill[a][0]].append(a)
            a+=1          

for s in sowdays:
    for i in range(len(cal)):
       sowdays[s].append(0)

# Leaves for particular sow for billable employees

for s in sowdays:
    for i in sow[s]:
        found = False
        if i in ybill:
            l=0
            while l <len(cal):
                if l<9:
                    d='0'+str(l+1)
                else:
                    d=str(l+1)
                if month <10:
                    m='0'+str(month)
                else:
                    m=str(month)
                df = str(d+'-'+str(m)+'-'+str(year))
                for j in range(len(ybill[i])):
                    if df >= bill[ybill[i][j]][4] and df <= bill[ybill[i][j]][5]:
                        s = bill[j][2]
                        break
                    j+=1
                if isinstance(Empid[i][l],int):
                    if Empid[i][l] > 0:
                        sowdays[s][l] = sowdays[s][l]
                    else:
                        sowdays[s][l] = sowdays[s][l] +1
                else:
                    if isinstance(Empid[i][l],str):
                        if Empid[i][l] == 'EL':
                            sowdays[s][l] = sowdays[s][l] +1
                l+=1
        else:
            l = 0
            while l < len(cal):
                val = sowdays[s][l]
                if isinstance(Empid[i][l],int):
                    if Empid[i][l] > 0:
                        sowdays[s][l] = val
                    else:
                        val+=1
                        sowdays[s][l]=val

                else:
                    if isinstance(Empid[i][l],str):
                        if Empid[i][l] == 'EL':
                            val+=1
                            sowdays[s][l]=val
                            
                l+=1

day = {}
for s in shadow:
    day[s] =[]
    i = 0
    while i<len(cal):
        if s in nbill:
            if i<9:
                d='0'+str(i+1)
            else:
                d=str(i+1)
            if month <10:
                m='0'+str(month)
            else:
                m=str(month)
            df = str(d+'-'+str(m)+'-'+str(year))
            for j in range(len(nbill[s])):
                if df >= nonbill[nbill[s][j]][4] and df <= nonbill[nbill[s][j]][5]:
                    x = nonbill[j][2]
                    print(x,i)
                    break
                j+=1
        else:
            x = Empsow[s]
        if x in sow:      
            if cal[i] == 0 and sowdays[x][i] ==0 and str(Empid[s][i]) > str(0) and str(Empid[s][i])!= 'EL':
                day[s].append(0)
            if cal[i] == 0 and sowdays[x][i] ==0 and str(Empid[s][i]) > str(0) and str(Empid[s][i])== 'EL':
                day[s].append('EL')
            elif cal[i]==0 and sowdays[x][i] !=0 and str(Empid[s][i]) >str(0) and str(Empid[s][i])!= 'EL':
                val = sowdays[Empsow[s]][i] -1
                day[s].append(Empid[s][i])
                sowdays[Empsow[s]][i] = val
            elif cal[i]==0 and sowdays[x][i] !=0 and (Empid[s][i] == 0 or Empid[s][i] == 'EL'):
                day[s].append('EL')
            elif cal[i]==0 and sowdays[x][i] ==0 and (Empid[s][i] == 0 or Empid[s][i] == 'EL'):
                day[s].append('EL')
            elif cal[i] == 1 and sowdays[x][i] == 0 and Empid[s][i] == 'WO':
                day[s].append('WO')
            elif cal[i] == 1 and sowdays[x][i] == 0 and Empid[s][i] > 0 :
                day[s].append(0)
            elif cal[i] == 1 and sowdays[x][i] !=0 and str(Empid[s][i]) >str(0) :
                val = sowdays[x][i] -1
                day[s].append(Empid[s][i])
                sowdays[x][i] = val
            elif cal[i] == 1 and sowdays[x][i] !=0 and Empid[s][i] == 'WO':
                day[s].append('WO')
        else:
            day[s].append(0)
        i+=1

non_day = {}
 
for d in day:
    i=0
    non_day[d] =[]
    while i < (len(cal)):
            if day[d][i] == 0 and cal[i]==0 and str(Empid[d][i]) > str(0) and str(Empid[d][i]) != 'EL':
                non_day[d].append(Empid[d][i])
            elif day[d][i] !=0 and cal[i] == 0 and str(Empid[d][i]) > str(0) and str(Empid[d][i]) != 'EL':
                non_day[d].append(0)
            elif Empid[d][i] == 'EL' or Empid[d][i] == 0:
                non_day[d].append('EL')
            elif day[d][i] !=0 and cal[i] == 1 and isinstance(Empid[d][i],int):
                non_day[d].append(0)
            elif day[d][i] == 'WO' and cal[i] == 1 and Empid[d][i] == 'WO':
                non_day[d].append('WO')
            elif day[d][i] ==0 and cal[i] == 1 and isinstance(Empid[d][i],int):
                non_day[d].append(Empid[d][i])
            elif day[d][i] == 0 and cal[i] == 1 and Empid[d][i] == 'WO':
                non_day[d].append('WO')                
            else:
                non_day[d].append('check :')
            i+=1
 
year = user_date.year
month = user_date.month
start_date = datetime(year, month, 1)
end_date = datetime(year, month, 1) + timedelta(days=32)
end_date = end_date.replace(day=1) - timedelta(days=1)

num_days = (end_date - start_date).days + 1

date_range = [start_date + timedelta(days=i) for i in range(num_days)]

df = pd.DataFrame(columns= ['EMP ID'] + ['EMP Name'] + ['SOW'] + ['Bilability Status'] + ['Bill Rate'] + ['Over Time Bill Rate'] + [date.strftime('%d-%b-%y %a') for date in date_range] + ['Total Hours'] + ['Total Amount'])

while True:   # repeat until the try statement succeeds
    try:
        path = input('Please, Enter the path where you want report to be generated: ')
        excel_writer = pd.ExcelWriter(path+'\invoice.xlsx',engine='openpyxl')
        break                             # exit the loop
    except IOError:
        input("Could not open file or incorrect path given ! Please close Report Excel or give the correct. Press Enter to retry.")
        # restart the loop

df.to_excel(excel_writer, sheet_name='Invoice', index = False)

wb = excel_writer.book
ws = excel_writer.sheets["Invoice"]
no_COMd = []

blue = PatternFill(start_color='ADD8E6',end_color='ADD8E6', fill_type='solid')
black = PatternFill(start_color='000000',end_color='000000', fill_type='solid')
yellow = PatternFill(start_color='FFEE99',end_color='FFEE99', fill_type='solid')
white=Font(color='FFFFFF',bold=True)

for cell in ws['1:1']:
    cell.fill=blue
    cell.font=Font(color='000000',bold=True)
    cell.alignment=Alignment(horizontal='center',vertical='center',text_rotation=90)

fmt = ['A','B','C','D','E','F'] 

for col in ws.columns:
    if col[0].column_letter in fmt:
        for cell in col:
            cell.fill = black
            cell.font = white
            cell.alignment = Alignment(horizontal='center',vertical='center')

for cell in ws['1:1']:
    if cell.value == 'Total Amount' or cell.value == 'Total Hours':
        cell.fill = black
        cell.font = white
        cell.alignment = Alignment(horizontal='center',vertical='center')
    
row_data=2

for e in Empid:
    if e in allid:
        col_data=1
        ws.cell(row=row_data, column= col_data, value = e)
        col_data = 2
        ws.cell(row=row_data, column= col_data, value = name[e])
        col_data = 3
        ws.cell(row=row_data, column= col_data, value = Empsow[e])
        col_data = 4
        ws.cell(row=row_data, column= col_data, value = status[e])
        col_data = 5
        ws.cell(row=row_data, column= col_data, value = Rate[e])
        col_data=7
        s=0
        if e in day:
                col_data=1
                ws.cell(row=row_data, column= col_data, value = e)
                col_data = 2
                ws.cell(row=row_data, column= col_data, value = name[e])
                col_data = 3
                ws.cell(row=row_data, column= col_data, value = Empsow[e])
                col_data = 4
                ws.cell(row=row_data, column= col_data, value = 'Shadow - Billable')
                col_data = 5
                ws.cell(row=row_data, column= col_data, value = Rate[e])
                col_data=7
                for d in day[e]:
                        ws.cell(row=row_data, column= col_data, value = d)
                        if isinstance(d,int):
                            s+=d
                        col_data+=1
                ws.cell(row=row_data, column= col_data, value = s)
                col_data+=1
                ws.cell(row=row_data, column= col_data, value = (s*Rate[e]))
                row_data+=1
                col_data=1
                ws.cell(row=row_data, column= col_data, value = e)
                col_data = 2
                ws.cell(row=row_data, column= col_data, value = name[e])
                col_data = 3
                ws.cell(row=row_data, column= col_data, value = Empsow[e])
                col_data = 4
                ws.cell(row=row_data, column= col_data, value = 'Non-billable')
                col_data = 5
                ws.cell(row=row_data, column= col_data, value = Rate[e])
                col_data=7
                s=0
                for d in non_day[e]:
                        ws.cell(row=row_data, column= col_data, value = d)
                        if isinstance(d,int):
                            s+=d
                        col_data+=1
                ws.cell(row=row_data, column= col_data, value = s)
                col_data+=1
                ws.cell(row=row_data, column= col_data, value = (s*Rate[e]))
                row_data+=1
        else:
            s=0
            for d in Empid[e]:
                    ws.cell(row=row_data, column= col_data, value = d)
                    if isinstance(d,int):
                        s+=d
                    col_data+=1
            ws.cell(row=row_data, column= col_data, value = s)
            col_data+=1
            ws.cell(row=row_data, column= col_data, value = (s*Rate[e]))
            row_data+=1
            s = 0
            if e in o_Empid:
                    col_data=1
                    ws.cell(row=row_data, column= col_data, value = e)
                    col_data = 2
                    ws.cell(row=row_data, column= col_data, value = name[e])
                    col_data = 3
                    ws.cell(row=row_data, column= col_data, value = Empsow[e])
                    col_data = 4
                    ws.cell(row=row_data, column= col_data, value = 'Overtime')
                    col_data = 6
                    ws.cell(row=row_data, column= col_data, value = (1.25 * Rate[e]))
                    col_data=7
                    for d in o_Empid[e]:
                            ws.cell(row=row_data, column= col_data, value = d)
                            if isinstance(d,int):
                                s+=d
                            col_data+=1
                    ws.cell(row=row_data, column= col_data, value = s)
                    col_data+=1
                    ws.cell(row=row_data, column= col_data, value = (s*(1.25*Rate[e])))
                    row_data+=1
    else:
        no_COMd.append(e)
        allid.append(e)

for row in ws.iter_rows(min_row=1):
    for cell in row:
        if cell.value == 'WO':
            cell.fill = yellow
            cell.alignment = Alignment(horizontal='center',vertical='center')
#ws.cell(row=row_data, column= col_data, value = data)

df2 = pd.DataFrame(columns= ['EMP ID'] + ['REASON'])

df2.to_excel(excel_writer, sheet_name='Defaulter', index = False)

wb = excel_writer.book
ws = excel_writer.sheets["Defaulter"]

nodate_id = []
for id in visited:
    if id not in nodate and id not in no_COMd:
        allid.remove(id)
    else:
        nodate_id.append(id)

row_data=2

for id in allid:
    ws.cell(row=row_data, column= 1, value = id)
    if id in no_COMd:
        ws.cell(row=row_data, column= 2, value = 'Employee Record not Present in COM')
    elif id in nodate_id:
        ws.cell(row=row_data, column= 2, value = '1 or more Missing Timesheets') 
    else:
        ws.cell(row=row_data, column= 2, value = 'No Timesheet Present') 
    row_data+=1
    i+=1

excel_writer._save()

if nofile:
    print(' ')
    print("*************************************************************")
    print('Following files were not considered as part of this report: ' )
    print("*************************************************************")
    print(' ')
    for i in nofile:
        print(i)

print("Done!!")
